import streamlit as st
import db  # Naka-link sa iyong db.py module
from datetime import datetime
import io

def render_excel_preview_module():
    """
    Main entry point ng module. 
    Pumapalit sa open_excel_preview_modal ng Tkinter para sa web rendering.
    """
    st.markdown("## 📊 POW Print & Excel Export Suite")
    st.caption("Pumili ng proyekto para makita ang formal document format layout at i-download ang structural openpyxl file nito.")

    # ==========================================================================
    # DETALYE 1: PAGKUHA NG MGA PROYEKTO SA DATABASE
    # ==========================================================================
    project_list = db.get_project_list()  # Hinahatak ang (pow_id, name, location)

    if not project_list:
        st.info("🗹 Walang mahanap na aktibong proyekto sa Cloud Database. Gumawa muna sa POW form window.")
        return

    # Ayusin ang format ng dropdown selectbox options
    project_options = {f"ID: {proj[0]} | {proj[1]}": proj[0] for proj in project_list}
    selected_option = st.selectbox("Pumili ng Proyekto sa Listahan:", options=list(project_options.keys()))
    
    # Kunin ang active target primary key id
    pow_id = project_options[selected_option]

    # Hilaan ng structural variables mula sa core backend module functions
    proj_info = db.get_project_details(pow_id)
    pow_items = db.get_items_by_project(pow_id)

    if not proj_info:
        st.error("❌ Error: Hindi mahanap ang detalye ng proyektong ito sa database.")
        return

    project_name, location = proj_info[0], proj_info[1]

    # ==========================================================================
    # DETALYE 2 & 3: FORMAL PRINT LAYOUT PREVIEW GENERATOR
    # ==========================================================================
    lines = []
    lines.append(f"{'':<25}Republic of the Philippines")
    lines.append(f"{'':<22}PROVINCE   OF   NUEVA   ECIJA")
    lines.append(f"{'':<26}Palayan City")
    lines.append(f"{'':<15}PROVINCIAL GENERAL SERVICES OFFICE")
    lines.append("")
    lines.append(f"{'':<27}PROGRAM OF WORKS")
    lines.append("")
    lines.append(f"Project:  {project_name}")
    lines.append(f"Location: {location}")
    lines.append("")
    
    lines.append("=" * 95)
    lines.append(f"{'ITEM':<8}{'QTY':<8}{'UNIT':<10}{'DESCRIPTION':<35}{'UNIT PRICE':<16}{'AMOUNT':<15}")
    lines.append("=" * 95)

    grand_total = 0.0
    for idx, item in enumerate(pow_items, start=1):
        qty = float(item[0])
        unit = item[1]
        raw_name = item[2]
        raw_price = item[3]
        
        name = str(raw_name).replace("\n", " ").replace("\r", " ").strip()
        try:
            price = float(raw_price) if raw_price is not None else 0.0
        except ValueError:
            price = 0.0

        amount = qty * price
        grand_total += amount
        
        short_name = name[:32] + "..." if len(name) > 32 else name
        lines.append(f"{idx:<8}{qty:<8.2f}{unit:<10}{short_name:<35}{price:>12,.2f}      {amount:>12,.2f}")

    lines.append("-" * 95)
    lines.append(f"{'TOTAL':<61}P     {grand_total:>22,.2f}")
    lines.append("=" * 95)
    lines.append("")

    lines.append(f"Prepared by:{'':<45}Checked by:")
    lines.append("")
    lines.append(f"       JONATHAN G. LADIGNON{'':<37}BENJAMIN N. RAMOS JR")
    lines.append(f"       Admin. Officer III  {'':<37}Engineer II")
    lines.append("")
    lines.append(f"Noted by:{'':<48}Recommending Approval:")
    lines.append("")
    lines.append(f"MARIO T. MARIANO{'':<41}ENGR. FLORECIO M. VALINO")
    lines.append("")
    lines.append(f"{'':<50}Approved:")
    lines.append("")
    lines.append(f"{'':<45}HON. AURELIO M. UMALI")
    lines.append(f"{'':<50}Governor")

    # Pagsasama-sama ng text block buffer pipeline string
    preview_string = "\n".join(lines)

    # I-render ang Preview panel layout gamit ang monospaced viewport
    st.markdown("### 📝 Layout Print Preview")
    st.code(preview_string, language="text", wrap_lines=False)

    # ==========================================================================
    # DETALYE 4: OPENPYXL MEMORY BUFFER LOGIC FOR SECURE DOWNLOAD
    # ==========================================================================
    def build_excel_memory_buffer():
        """Bumubuo ng openpyxl spreadsheet object direkta sa isang volatile RAM stream."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Program of Work"
        ws.views.sheetView[0].showGridLines = True

        # Page setup rules configuration profiles
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_LEGAL
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1   
        ws.page_setup.fitToHeight = 0  

        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75

        font_title = Font(name="Arial", size=10, bold=True)
        font_regular = Font(name="Arial", size=10)
        font_bold_body = Font(name="Arial", size=10, bold=True)
        
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        ws['A1'] = "Republic of the Philippines"
        ws['A2'] = "PROVINCE   OF   NUEVA   ECIJA"
        ws['A3'] = "Palayan City"
        ws['A4'] = "PROVINCIAL GENERAL SERVICES OFFICE"
        ws['A6'] = "PROGRAM OF WORKS"
        
        for r in range(1, 7):
            if ws.cell(row=r, column=1).value:
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
                header_cell = ws.cell(row=r, column=1)
                header_cell.alignment = align_center
                header_cell.font = font_title

        ws['A7'] = f"Project: {project_name}"
        ws['A8'] = f"Location: {location}"
        ws['A7'].font = font_bold_body
        ws['A8'].font = font_bold_body

        headers_list = ["ITEM", "QTY", "UNIT", "DESCRIPTION", "UNIT PRICE", "AMOUNT"]
        for col_num, header_text in enumerate(headers_list, start=1):
            c_cell = ws.cell(row=9, column=col_num, value=header_text)
            c_cell.font = font_bold_body
            c_cell.alignment = align_center
            
        row_tracker = 10
        computed_total = 0.0
        
        for idx, item in enumerate(pow_items, start=1):
            qty = float(item[0])
            unit = item[1]
            raw_name = item[2]
            raw_price = item[3]
            
            name = str(raw_name).replace("\n", " ").replace("\r", " ").strip()
            try:
                price = float(raw_price) if raw_price is not None else 0.0
            except ValueError:
                price = 0.0

            amount = qty * price
            computed_total += amount

            ws.cell(row=row_tracker, column=1, value=idx).alignment = align_center
            ws.cell(row=row_tracker, column=2, value=qty).alignment = align_center
            ws.cell(row=row_tracker, column=3, value=unit).alignment = align_center
            ws.cell(row=row_tracker, column=4, value=name).alignment = align_left
            
            p_cell = ws.cell(row=row_tracker, column=5, value=price)
            p_cell.number_format = '#,##0.00'
            p_cell.alignment = align_right
            
            a_cell = ws.cell(row=row_tracker, column=6, value=amount)
            a_cell.number_format = '#,##0.00'
            a_cell.alignment = align_right

            for column_pos in range(1, 7):
                ws.cell(row=row_tracker, column=column_pos).font = font_regular
                
            row_tracker += 1

        row_tracker += 1 
        ws.cell(row=row_tracker, column=5, value="Total  P").font = font_bold_body
        ws.cell(row=row_tracker, column=5).alignment = Alignment(horizontal="right", vertical="center")
        
        f_total_cell = ws.cell(row=row_tracker, column=6, value=computed_total)
        f_total_cell.font = font_bold_body
        f_total_cell.number_format = '"P" #,##0.00'
        f_total_cell.alignment = align_right

        double_bottom_border = Border(bottom=Side(style='double'))
        last_item_amount_cell = ws.cell(row=row_tracker - 2, column=6)
        last_item_amount_cell.border = double_bottom_border

        row_tracker += 2
        ws.cell(row=row_tracker, column=1, value="Prepared by:                                                                                Checked by:").font = font_regular
        
        row_tracker += 2
        ws.cell(row=row_tracker, column=1, value="        JONATHAN G. LADIGNON                                                                BENJAMIN N. RAMOS JR.").font = font_bold_body
        
        row_tracker += 1
        ws.cell(row=row_tracker, column=1, value="             Admin. Officer III                                                                                 Engineer II").font = font_regular

        row_tracker += 2
        ws.cell(row=row_tracker, column=1, value="Noted by:                                                                                      Recommending Approval:").font = font_regular

        row_tracker += 2
        ws.cell(row=row_tracker, column=1, value="        MARIO T. MARIANO                                                                     ENGR. FLORECIO M. VALINO").font = font_bold_body
        
        row_tracker += 1
        ws.cell(row=row_tracker, column=1, value="             Engineer IV                                                                                        PGS-Officer").font = font_regular

        row_tracker += 2
        ws.cell(row=row_tracker, column=4, value="                             Approved:").font = font_regular

        row_tracker += 2
        ws.cell(row=row_tracker, column=4, value="                     HON. AURELIO M. UMALI").font = font_bold_body
        row_tracker += 1
        ws.cell(row=row_tracker, column=4, value="                                   Governor").font = font_regular

        column_widths = {'A': 4.57, 'B': 4.86, 'C': 7.00, 'D': 47.14, 'E': 11.57, 'F': 14.57}
        for col_letter, width_size in column_widths.items():
            ws.column_dimensions[col_letter].width = width_size

        ws.row_dimensions[9].height = 20
        for r_idx in range(10, row_tracker + 15):
            ws.row_dimensions[r_idx].height = 16

        if len(pow_items) > 35:
            from openpyxl.worksheet.pagebreak import Break
            ws.row_breaks.append(Break(id=45))

        # I-serialize ang file papunta sa io.BytesIO memory stream stream channel
        excel_stream = io.BytesIO()
        wb.save(excel_stream)
        excel_stream.seek(0)
        return excel_stream.getvalue()

    # ==========================================================================
    # ACTION CONTROL BAR (DOWNLOAD HANDLERS)
    # ==========================================================================
    st.write("---")
    
    # Linisin ang pangalan para maging ligtas na file string
    clean_filename = project_name.replace(' ', '_').replace('/', '-').replace('\\', '-')
    final_output_name = f"POW_{clean_filename}.xlsx"

    try:
        # Kunin ang data array mula sa functional memory generator hook natin
        excel_data_binary = build_excel_memory_buffer()
        
        st.success(f"📌 Handa na ang Openpyxl Layout engine para sa Proyektong: **{project_name}**")
        
        # Native web application layout file downloader tool component
        st.download_button(
            label="📥 EXPORT TO ACTUAL EXCEL FILE",
            data=excel_data_binary,
            file_name=final_output_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True
        )
    except Exception as ex:
        st.error(f"❌ Structural build runtime exception occurred: {ex}")
