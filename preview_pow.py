import streamlit as st
import pandas as pd
import db  # Naka-link sa iyong db.py module
import os
import subprocess
from openpyxl import load_workbook

# Import natin ang nakaraang module para sa Excel Downloader Engine
import preview_module  

def render_preview_pow_module():
    st.markdown("## 📋 PREVIEW SAVED PROGRAM OF WORK (POW)")
    st.caption("Pamahalaan, i-edit, o burahin ang mga naka-save na structural projects sa database system.")

    # SIGURADUHING MAY STATE TRACKING SA MGA INPUT ROWS
    if "editing_pow_id" not in st.session_state:
        st.session_state.editing_pow_id = None

    # ==========================================================================
    # KALIWA/ITAAS: PROJECT SELECTOR & ACTION GATEWAY
    # ==========================================================================
    projects = db.get_project_list()
    
    if not projects:
        st.info("🗹 Walang mahanap na aktibong proyekto sa database. Gumawa muna ng bago.")
        return

    # I-map ang projects para sa malinis na interface array
    project_options = {f"ID: {p[0]} | {p[1]}": {"id": p[0], "name": p[1], "location": p[2]} for p in projects}
    selected_label = st.selectbox("🎯 Pumili ng Proyekto sa Listahan:", options=list(project_options.keys()))
    
    current_proj = project_options[selected_label]
    pow_id = current_proj["id"]
    project_name = current_proj["name"]
    location = current_proj["location"]

    # Ipakita ang aktibong Lokasyon
    st.markdown(f"**📍 Lokasyon:** `{location}`")

    # ==========================================================================
    # KANAN: ITEMS DISPLAY LIST TABLE
    # ==========================================================================
    associated_items = db.get_items_by_project(pow_id)
    
    table_data = []
    grand_total = 0.0
    
    for idx, item in enumerate(associated_items, start=1):
        qty = float(item[0])
        unit = item[1]
        name = item[2]
        price = float(item[3])
        total = qty * price
        grand_total += total
        
        table_data.append({
            "#": idx,
            "Qty": qty,
            "Unit": unit,
            "Item Description": name,
            "Unit Price (₱)": price,
            "Total Price (₱)": total
        })

    # I-render bilang malinis na DataFrame
    if table_data:
        df = pd.DataFrame(table_data)
        st.dataframe(
            df.style.format({"Qty": "{:.2f}", "Unit Price (₱)": "{:,.2f}", "Total Price (₱)": "{:,.2f}"}),
            use_container_width=True,
            hide_index=True
        )
    else:
        st.warning("⚠️ Walang laman na mga aytem ang proyektong ito.")

    # 💰 SUMMARY BAR CONTROLLER
    st.metric(label="PROJECT TOTAL COST", value=f"₱ {grand_total:,.2f}")

    # ==========================================================================
    # BOTTOM ACTION CONTROL BOARD PANELS
    # ==========================================================================
    col1, col2, col3 = st.columns(3)

    with col1:
        # Paggamit ng functional script mula sa kabilang sheet module file natin kanina
        if st.button("👁️ Open Layout Preview & Download Suite", use_container_width=True, type="primary"):
            st.info("I-on ang module tab sa ibaba para makita ang Print Preview form.")
            # Trigger switcher pattern logic
            st.session_state.page_view_mode = "print_preview" 

    with col2:
        if st.button("✏️ Edit POW Record / Update Items", use_container_width=True):
            st.session_state.editing_pow_id = pow_id
            trigger_edit_modal_dialog(pow_id, project_name, location, associated_items)

    with col3:
        if st.button("❌ Delete Entire POW", use_container_width=True, type="secondary"):
            # Gumamit ng session state toggle para sa delete safety verification trigger guard
            st.session_state.confirm_delete_id = pow_id

    # CONTROL FOR SEPARATE TRIGGER CONFIRMATION SYSTEM
    if "confirm_delete_id" in st.session_state and st.session_state.confirm_delete_id == pow_id:
        st.error(f"⚠️ **KUMPIRMASYON:** Sigurado ka bang buburahin ang buong proyekto: **{project_name}**? Hindi na ito mababawi.")
        c_del1, c_del2 = st.columns(2)
        with c_del1:
            if st.button("Oo, Burahin ang Lahat", type="primary", use_container_width=True):
                if db.delete_pow_from_sql(pow_id):
                    st.success("Matagumpay na nabura ang buong POW record.")
                    del st.session_state.confirm_delete_id
                    st.rerun()
        with c_del2:
            if st.button("I-cancel", use_container_width=True):
                del st.session_state.confirm_delete_id
                st.rerun()

# ==============================================================================
# SUB-MODULE DIALOG: POPUP EDIT MODE SYSTEM (openpyxl & SQL SYNC)
# ==============================================================================
@st.dialog("✏️ EDIT MODE - Update POW Record & Sync Pipelines", width="large")
def trigger_edit_modal_dialog(pow_id, current_name, current_location, associated_items):
    """Pumapalit sa Toplevel modal frame engine ng Tkinter UI window block."""
    
    st.markdown("### 🏢 Details of POW")
    new_name = st.text_input("Project Title / Name:", value=current_name)
    new_loc = st.text_input("Project Location:", value=current_location)

    st.markdown("### 📊 List of Items (Inline Spreadsheet Data Editor)")
    st.caption("💡 Pwede mong palitan direkta ang values sa table sa ibaba, magdagdag sa pinakahuling linya, o pindutin ang `Delete` button.")

    # Ihanda ang dataset para sa dynamic interactive editor dataframe grid
    edit_rows = []
    for item in associated_items:
        edit_rows.append({
            "QTY": float(item[0]),
            "UNIT": str(item[1]),
            "ITEM DESCRIPTION": str(item[2]),
            "UNIT PRICE": float(item[3]),
            "ORIGINAL NAME": str(item[2])  # Nakatagong identifier node para sa Excel tracking link
        })
    
    df_editable = pd.DataFrame(edit_rows)

    # NATIVE STREAMLIT DATA EDITOR SYSTEM (Lunas sa mahabang manual row configuration forms!)
    edited_df = st.data_editor(
        df_editable,
        num_rows="dynamic", # Pinapayagan ang ➕ Add Line at pagbura natively sa browser grid
        use_container_width=True,
        column_config={
            "ORIGINAL NAME": st.column_config.TextColumn(help="Hidden system identity map tracker.", disabled=True)
        }
    )

    st.write("---")
    
    # --- SAVE ACTION EXECUTION CORE ENGINE ---
    if st.button("💾 SAVE ALL CHANGES & OVERWRITE DATABASE", type="primary", use_container_width=True):
        new_name_clean = new_name.strip().title()
        new_loc_clean = new_loc.strip().title()

        if not new_name_clean or not new_loc_clean:
            st.error("❌ Huwag iwanang blangko ang Name at Location, boss.")
            return

        excel_path = r"G:\jrm\master_items.xlsx"
        
        # --- EXCEL SPREADSHEET MASTER FILE SYSTEM CHECK & LOGIC ---
        if not os.path.exists(excel_path):
            st.error(f"❌ File Error: Hindi mahanap ang Excel file sa: {excel_path}. Siguraduhing naka-mount ang G: drive network module node.")
            return

        try:
            wb = load_workbook(excel_path)
            ws = wb.active
            
            # Repopulate final matrix array stack maps
            final_items_to_save = []
            
            for index, row in edited_df.iterrows():
                try:
                    q = float(row["QTY"])
                    u = str(row["UNIT"]).upper()
                    d = str(row["ITEM DESCRIPTION"]).strip()
                    p = float(row["UNIT PRICE"])
                    
                    # Hawakan ang safe execution para sa mga bagong idinagdag na row blocks
                    orig_d = str(row["ORIGINAL NAME"]).strip() if pd.notna(row["ORIGINAL NAME"]) else d.strip()
                    
                    final_items_to_save.append((q, u, d, p))

                    # Mag-loop sa Excel para hanapin at palitan ang tumutugmang master item
                    target_row = None
                    for ex_row in range(2, ws.max_row + 1):
                        cell_val = ws.cell(row=ex_row, column=1).value
                        if cell_val and str(cell_val).strip().lower() == orig_d.lower():
                            target_row = ex_row
                            break

                    if target_row:
                        ws.cell(row=target_row, column=1, value=d)
                        ws.cell(row=target_row, column=2, value=u)
                        ws.cell(row=target_row, column=3, value=p)
                    else:
                        new_row = ws.max_row + 1
                        ws.cell(row=new_row, column=1, value=d)
                        ws.cell(row=new_row, column=2, value=u)
                        ws.cell(row=new_row, column=3, value=p)
                        
                except Exception as row_err:
                    st.error(f"Format Conversion Value Error on line item index: {index} | {row_err}")
                    wb.close()
                    return

            # I-save ang Workbook structural updates
            wb.save(excel_path)
            wb.close()
            
        except Exception as e:
            st.error(f"❌ Excel Open/Save Pipeline Failure Mode: {e}")
            return

        # --- SUBPROCESS PYTHON EXECUTION AGENT TRIGGER ---
        try:
            subprocess.run(["py", "import_master.py"], cwd=r"G:\jrm", check=True, capture_output=True, text=True)
        except Exception as err:
            st.warning(f"⚠️ Na-save sa Excel pero nabigong patakbuhin ang 'import_master.py' monitoring script core.\n\nDetalye: {err}")

        # --- REVISED SQL SERVER METADATA SYNC CONTROLLER ---
        success_main = db.update_project_main_details(pow_id, new_name_clean, new_loc_clean)
        success_items = db.update_project_items_batch(pow_id, final_items_to_save)

        if success_main and success_items:
            st.success("🎉 Swabe ang ikot, boss! Na-save sa Excel, SQL Master at Cloud Server Core! Nagre-refresh...")
            st.rerun()
        else:
            st.error("❌ SQL Synchronization Exception Error: May sumabog sa batch transaction data tables.")
