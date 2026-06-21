import os
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import db  # Siniguradong naka-import ang database module mo

# --- SESSION STATE INITIALIZATION ---
if "temporary_items" not in st.session_state:
    st.session_state.temporary_items = []
if "show_finalize_form" not in st.session_state:
    st.session_state.show_finalize_form = False

def render_add_pow_module():
    st.markdown("## ➕ CREATE PROGRAM OF WORK (POW)")
    st.caption("Workspace Module: Local Input System Engine for POW Generation")
    st.divider()

    # --- ITEM ENTRY FORM COMPONENT ---
    st.markdown("#### 📦 Item Entry (Max 150 Items)")
    
    with st.container(border=True):
        col_qty, col_unit, col_search = st.columns([1, 1.5, 4.5])
        
        with col_qty:
            qty_input = st.number_input("Qty:", min_value=0, value=0, step=1, key="entry_qty")
            
        with col_unit:
            unit_choices = ["PC", "GAL", "LIT", "TIN", "BOX", "BAG", "CU.M"]
            unit_input = st.selectbox("Unit:", options=unit_choices, index=0, key="entry_unit")
            
        with col_search:
            # Query at default search filter logic para sa dropdown autocomplete capability
            search_query = st.text_input("Type to search item from database... (Press Enter to list choices)", key="item_search_field")
            
            # Kumukuha ng options base sa tinype ng user, o nagpapakita ng top basic results
            db_results = db.search_master_items(search_query) if search_query else []
            
            # Gumawa ng readable label string mapping para sa dropdown select box
            options_map = {}
            options_list = ["-- Pumili ng Aytem sa Listahan / Magrehistro ng Bago --"]
            
            for item in db_results:
                # assuming db returns: [name, unit, price]
                name, unit, price = item[0], item[1], float(item[2])
                label_string = f"🏷️ {name} ({unit}) - ₱{price:,.2f}"
                options_list.append(label_string)
                options_map[label_string] = {"name": name, "unit": unit, "price": price}
                
            selected_option = st.selectbox("Select Item Matched:", options=options_list, index=0)

        # Control fields para sa dynamic price override at manual text injection kung bagong rehistro
        st.write("")
        col_name_input, col_price_input, col_btn = st.columns([4, 2, 2])
        
        with col_name_input:
            # Awtomatikong pinupunan kung may napili sa dropdown list, pwedeng i-type kung bago
            default_name = ""
            if selected_option != options_list[0]:
                default_name = options_map[selected_option]["name"]
            else:
                default_name = search_query
                
            final_item_name = st.text_input("Final Item Description Name:", value=default_name, key="final_name")
            
        with col_price_input:
            default_price = 0.00
            if selected_option != options_list[0]:
                default_price = options_map[selected_option]["price"]
                
            final_unit_price = st.number_input("Unit Price (₱):", min_value=0.00, value=default_price, step=0.01, format="%.2f", key="final_price")

        with col_btn:
            st.write("<div style='padding-top:24px;'></div>", unsafe_allow_html=True) # visual spacer alignment
            add_submitted = st.button("➕ Add to List", type="primary", use_container_width=True)

    # --- ACTION EXECUTION FOR ADDING ITEM ---
    if add_submitted:
        name_clean = final_item_name.strip().title()
        unit_clean = unit_input.strip().upper()
        
        if not name_clean:
            st.warning("⚠️ Input Error: Paki-lagay ang Item Name bago magpatuloy.")
        elif len(st.session_state.temporary_items) >= 150:
            st.error("🚨 Limit Reached: Hanggang 150 items lamang ang pwedeng ilagay sa isang POW.")
        else:
            # Database sync checkpoint: Suriin kung bago ang aytem sa listahan
            check_exists = db.search_master_items(name_clean)
            is_new_item = True
            for db_item in check_exists:
                if db_item[0].lower() == name_clean.lower():
                    is_new_item = False
                    break
            
            # Isulat sa database master table at excel kung ito ay bago
            if is_new_item:
                if hasattr(db, 'add_new_master_item'):
                    db.add_new_master_item(name_clean, unit_clean, final_unit_price)
                
                excel_path = r"G:\jrm\master_items.xlsx"
                if os.path.exists(excel_path):
                    try:
                        wb = load_workbook(excel_path)
                        ws = wb.active
                        new_row = ws.max_row + 1
                        ws.cell(row=new_row, column=1, value=name_clean)
                        ws.cell(row=new_row, column=2, value=unit_clean)
                        ws.cell(row=new_row, column=3, value=final_unit_price)
                        wb.save(excel_path)
                        wb.close()
                    except Exception as e:
                        st.sidebar.warning(f"[Excel Sync Warning] {e}")
            
            # I-push ang record dictionary sa internal session arrays memory storage
            st.session_state.temporary_items.append({
                'Qty': int(qty_input),
                'Unit': unit_clean,
                'Item Name': name_clean,
                'Unit Price': float(final_unit_price),
                'Total Price': int(qty_input) * float(final_unit_price)
            })
            st.toast(f"Added: {name_clean} to temporary list buffering layer.", icon="✅")
            st.rerun()

    # --- TABLE PREVIEW GRID CONTROL LAYER ---
    st.markdown(f"#### 📊 Current Items Added Block Grid ({len(st.session_state.temporary_items)} / 150 items)")
    
    if st.session_state.temporary_items:
        # I-convert ang raw records arrays list pabalik sa readable panda dataframe table view
        df_preview = pd.DataFrame(st.session_state.temporary_items)
        
        # Dagdagan ng index column row indicator framework counter matrix
        df_preview.index = df_preview.index + 1
        
        # Gamitin ang st.data_editor para interactive at madaling ma-edit o mabura ang maling entries
        st.caption("💡 Tip: Pwede mong direktang baguhin o burahin ang rows sa editable table grid panel sa ibaba.")
        edited_items_df = st.data_editor(df_preview, use_container_width=True, num_rows="dynamic", key="pow_table_editor")
        
        # Pagpapanatili ng real-time state synchronization kapag may pagbabago sa grid matrix data lines
        # (Sinasakop na nito ang functionality ng dating remove_selected_item runtime routing)
        if st.button("🔄 Sync Changes / Recalculate Table Grid Data", type="secondary"):
            updated_list = []
            for _, row in edited_items_df.iterrows():
                updated_list.append({
                    'Qty': int(row['Qty']),
                    'Unit': str(row['Unit']).upper(),
                    'Item Name': str(row['Item Name']),
                    'Unit Price': float(row['Unit Price']),
                    'Total Price': int(row['Qty']) * float(row['Unit Price'])
                })
            st.session_state.temporary_items = updated_list
            st.rerun()

        # --- CALCULATE TOTALS DISPLAY PANEL ---
        grand_total = sum(item['Total Price'] for item in st.session_state.temporary_items)
        
        st.markdown(
            f"<div style='text-align: right; background-color: #edf2f7; padding: 15px; border-radius: 5px; margin-top: 10px;'>"
            f"<h3 style='margin:0; color: #2b6cb0;'>GRAND TOTAL: ₱ {grand_total:,.2f}</h3>"
            f"</div>", 
            unsafe_allow_html=True
        )

        # Clear All Button to refresh buffer memory pipeline routing sequence
        col_spacer, col_action_save = st.columns([3, 1])
        with col_action_save:
            if st.button("💾 SAVE WHOLE POW MODULE STRUCTURE", type="primary", use_container_width=True):
                st.session_state.show_finalize_form = True

        # --- FINALIZE POW MODAL SECTION REPLACEMENT CONTROLLER ---
        if st.session_state.show_finalize_form:
            st.divider()
            with st.form("finalize_pow_metadata_form"):
                st.markdown("#### 🏛️ Finalize Project Administrative Details")
                st.info("Ipasok ang Pangwakas na Detalye para sa pagsusulat ng record sequence node block.")
                
                proj_name_input = st.text_input("Project Name / Title Descriptor:")
                location_input = st.text_input("Project Exact Geographic Location:")
                
                col_c1, col_c2 = st.columns(2)
                with col_c1:
                    btn_save_submit = st.form_submit_button("Confirm & Save to MySQL Server Core", use_container_width=True)
                with col_c2:
                    btn_cancel = st.form_submit_button("Cancel / Return Back", use_container_width=True)
                    
                if btn_cancel:
                    st.session_state.show_finalize_form = False
                    st.rerun()
                    
                if btn_save_submit:
                    p_name = proj_name_input.strip().title()
                    p_loc = location_input.strip().title()
                    
                    if not p_name or not p_loc:
                        st.error("❌ Kulang na Impormasyon: Ipasok ang Project Name at Location bago i-save.")
                    else:
                        # Streamlit architecture conversion for local relational schema injection block array
                        # Inihahanda ang data payload architecture na tinatanggap ng iyong db.save_pow_to_sql script
                        db_payload = []
                        for item in st.session_state.temporary_items:
                            db_payload.append({
                                'qty': item['Qty'],
                                'unit': item['Unit'],
                                'name': item['Item Name'],
                                'price': item['Unit Price']
                            })
                            
                        success = db.save_pow_to_sql(p_name, p_loc, db_payload)
                        if success:
                            st.success("🎉 Success: Matagumpay na nai-save ang Program of Work (POW) sa database layer system!")
                            st.session_state.temporary_items = []
                            st.session_state.show_finalize_form = False
                            st.rerun()
                        else:
                            st.error("🚨 Database Error: Nagka-error ang SQL server architecture engine sa pagproseso ng query.")
    else:
        st.info("💡 Walang laman ang temporary working table buffer block sa kasalukuyan. Magsimula sa pagdadagdag ng aytem sa form panel sa itaas.")

if __name__ == "__main__":
    render_add_pow_module()
